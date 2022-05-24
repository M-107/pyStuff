import os
import numpy as np
import tkinter as tk
import seaborn as sns
from scipy import sparse
from scipy.integrate import simps
from openpyxl import load_workbook
from scipy.signal import find_peaks
from matplotlib import pyplot as plt
from math import sin, cos, pi, radians
from scipy.sparse.linalg import spsolve
from tkinter.filedialog import askopenfilenames

testing = 1

def baseline_als(y, lam, p, niter=10):
  L = len(y)
  D = sparse.diags([1,-2,1],[0,-1,-2], shape=(L,L-2))
  w = np.ones(L)
  for i in range(niter):
    W = sparse.spdiags(w, 0, L, L)
    Z = W + lam * D.dot(D.transpose())
    z = spsolve(Z, w*y)
    w = p * (y > z) + (1-p) * (y < z)
  return z

def braggs(q):
  d = 0.154 / (2 * (sin(radians(q / 2)))) * 10
  return(d)

def scherrer(q, B):
    b = 0.114
    lam = 1.54
    k = 0.9
    bB = b/B
    betaB = 4.1871*(bB**6) - 12.979*(bB**5) + 14.206*(bB**4) - 5.4662*(bB**3) - 0.7782*(bB**2) - 0.1756*(bB) + 1.0026
    beta = B * betaB
    D = ((lam * k)  /(beta * cos((q * pi)/360))) * (360 / (2 * pi))
    return(D)

if testing == False:
  root = tk.Tk()
  root.withdraw()
  files = askopenfilenames(title="Select the Excel files with RTG data.", filetypes=[("Excel files", ".xlsx")])
elif testing == True:
  files = [r"---path to file---"]

for x, filename in enumerate(files, 1):
  print(f"Working on file {filename[-8:-5]} ({x}/{len(files)})")
  wb = load_workbook(filename)
  ws = wb.worksheets[0]
  number = ws["D1"].value
  name = ws["D2"].value
  image = f"{os.path.dirname(filename)}\{number}.png"
  year = ""
  if image.find("2001") != -1:
    year = "2001\n"
  if image.find("2021") != -1:
    year = "2021\n"

  x = [i.value for i in ws["A"] if i.value != None]
  y = [i.value for i in ws["B"] if i.value != None]
  xStart = min(range(len(x)), key=lambda i: abs(x[i]-10))
  xEnd = min(range(len(x)), key=lambda i: abs(x[i]-35.05))
  x = x[xStart:xEnd]
  y = y[xStart:xEnd]

  xMin = min(x)
  xMax = max(x)
  x = np.array(x).astype(float)
  y = np.array(y).astype(float)

  if image.find("KT-DAHQ") != -1:
    lamBase, pBase, lamAir, pAir, niter, width, height, prominence = (1000, 0.0001, 1000000000000, 0.0001, 100, (4, 10000), (150, 50000), 0.0001)
    known = [["I", [20, 23, 28, 30], "r--"], ["II", [18, 25], "g--"], ["KT", [17.5, 25, 27.8], "y--"], ["DAHQ", [17, 19.3, 21.7, 23.7, 26, 27.2, 28.7, 31.3], "y--"]]
  elif image.find("KIF-DAHQ") != -1:
    lamBase, pBase, lamAir, pAir, niter, width, height, prominence = (1000, 0.0001, 1000000, 0.00001, 100, (10, 10000), (500, 50000), 0.0001)
    known = [["I", [13, 18, 26], "r--"], ["II", [16, 23, 29], "g--"], ["KIF", [15.8, 18.3, 24.4, 27.8], "y--"], ["DAHQ", [17, 19.3, 21.7, 23.7, 26, 27.2, 28.7, 31.3], "y--"]]
  elif image.find("KIF-DARES") != -1:
    lamBase, pBase, lamAir, pAir, niter, width, height, prominence = (1000, 0.00001, 1000000, 0.00001, 100, (10, 1000000), (10, 50000), 0.000000000000001)
    known = [["I", [13, 18, 26], "r--"], ["KIF", [15.8, 18.3, 24.4, 27.8], "y--"]]
  elif image.find("KIF-DA3FED") != -1:
    lamBase, pBase, lamAir, pAir, niter, width, height, prominence = (1000, 0.0001, 1000000000000, 0.0001, 100, (10, 10000), (500, 50000), 0.0001)
    known = [["I", [17, 18], "r--"], ["KIF", [15.8, 18.3, 24.4, 27.8], "y--"], ["DA3FED", [9.4, 13.7, 16.9, 19.5, 23.5, 24.7, 25.7, 27.2], "y--"]]
  else:
    lamBase, pBase, lamAir, pAir, niter, width, height, prominence = (1000, 0.0001, 1000000000000, 0.0001, 100, (10, 10000), (500, 50000), 0.0001)
  
  cats = []
  pts = ["474", "542", "458", ]
  naac = ["439", "769", "785", "787", "473", "712", "768", "829", "810", "812", "767", "772", "801", "804", "789", "791", "799", "800"]
  for i in pts:
    if image.find(i) != -1:
      cats.append(["PTS", [16.2, 19, 24, 26.9], "y:"])
  for i in naac:
    if image.find(i) != -1:
      cats.append(["NaAc", [8.6, 22.8, 29.8], "y:"])

  base = baseline_als(y, lamBase, pBase, niter)
  air = baseline_als(base, lamAir, pAir, niter)

  areaAll = simps(y, dx=5)
  areaBase = simps(base, dx=5)
  areaAir = simps(air, dx=5)

  areaSample = areaAll - areaAir
  areaAmorph = areaBase - areaAir
  crystalRatio = 100 - (areaAmorph / areaSample * 100)

  peaks, pInfo = find_peaks(y[:(int(len(x)*7/8))]-base[:(int(len(x)*7/8))], width=width, height=height, prominence=prominence)

  sns.set_theme()
  sns.set_style("whitegrid")
  plt.plot(x, y, "b")
  plt.plot(x, base, "b--")
  plt.fill_between(x, air, color="blue", alpha=0.1)
  plt.plot(x[peaks], y[peaks],  "xr")
  yHop = 12
  # for i in known:
  #   tSize = len(i[0])
  #   for j in i[1]:
  #     yVal = y[min(range(len(x)), key=lambda i: abs(x[i]-j))]
  #     plt.plot((j, j), (0, yVal), i[2])
  #     plt.annotate(f"{i[0]}", (j, yVal), xytext=(+(1*tSize),-(3+yHop)), textcoords="offset points", fontfamily="DejaVu Sans Mono", alpha=0.5, fontsize=12)
  # for i in cats:
  #   tSize = len(i[0])
  #   for j in i[1]:
  #     yVal = y[min(range(len(x)), key=lambda i: abs(x[i]-j))]
  #     plt.plot((j, j), (0, yVal), i[2])
  #     plt.annotate(f"{i[0]}", (j, yVal), xytext=(-(1*tSize),-(3+yHop)), textcoords="offset points", fontfamily="DejaVu Sans Mono", alpha=0.5, fontsize=8)
  for n, i in enumerate(peaks):
          q = x[i]
          B = x[i+int(pInfo['widths'][n]/2)] - x[i-int(pInfo['widths'][n]/2)]
          xPos = str(round(x[i], 1))
          braggsVal = str(round(braggs(q), 1))
          scherrerVal = str(round(scherrer(q, B), 1))
          area = str(int(round(((pInfo['peak_heights'][n] * pInfo['widths'][n])), 0)))
          widthVal = str(round(pInfo['widths'][n], 0))
          widthValUnits = str(round(x[i+int(pInfo['widths'][n]/2)]-x[i-int(pInfo['widths'][n]/2)], 0))
          height = str(int(round(pInfo['peak_heights'][n], 0)))
          promi = str(round(pInfo['prominences'][n], 1))
          size = max([len(xPos), len(braggsVal), len(scherrerVal), len(area), len(widthVal), len(widthValUnits), len(height), len(promi)]) - 1
          # plt.annotate(f"X:{xPos:>{size}}", (x[i], y[i]), xytext=(5,-(3+1*yHop)), textcoords="offset points", fontfamily="DejaVu Sans Mono")
          # plt.annotate(f"B:{braggsVal:>{size}}", (x[i], y[i]), xytext=(5,-(3+2*yHop)), textcoords="offset points", fontfamily="DejaVu Sans Mono")
          # plt.annotate(f"S:{scherrerVal:>{size}}", (x[i], y[i]), xytext=(5,-(3+3*yHop)), textcoords="offset points", fontfamily="DejaVu Sans Mono")
          # plt.annotate(f"A:{area:>{size}}", (x[i], y[i]), xytext=(5,-(3+3*yHop)), textcoords="offset points", fontfamily="DejaVu Sans Mono")
          # plt.annotate(f"W:{widthVal:>{size}}", (x[i], y[i]), xytext=(5,-(3+2*yHop)), textcoords="offset points", fontfamily="DejaVu Sans Mono")
          # plt.annotate(f"W:{widthValUnits:>{size}}", (x[i], y[i]), xytext=(5,-(3+2*yHop)), textcoords="offset points", fontfamily="DejaVu Sans Mono")
          # plt.annotate(f"H:{height:>{size}}", (x[i], y[i]), xytext=(5,-(3+3*yHop)), textcoords="offset points", fontfamily="DejaVu Sans Mono")
          # plt.annotate(f"P:{promi:>{size}}", (x[i], y[i]), xytext=(5,-(3+4*yHop)), textcoords="offset points", fontfamily="DejaVu Sans Mono")
          # plt.plot((x[i], x[i]), (y[i], y[i]-pInfo["peak_heights"][n]))
          # plt.plot((x[i-int(pInfo['widths'][n]/2)], x[i+int(pInfo['widths'][n]/2)]), (y[i]-pInfo["peak_heights"][n]/2, y[i]-pInfo["peak_heights"][n]/2))
          plt.plot((x[i-int(pInfo['widths'][n])], x[i], x[i+int(pInfo['widths'][n])], x[i-int(pInfo['widths'][n])]), (y[i]-pInfo["peak_heights"][n], y[i], y[i]-pInfo["peak_heights"][n], y[i]-pInfo["peak_heights"][n]), alpha=0.8)
          # plt.plot((x[i], x[i+int(pInfo['widths'][n]/2)]), (y[i], y[i]-pInfo["peak_heights"][n]))
          # plt.plot((x[i-int(pInfo['widths'][n]/2)], x[i+int(pInfo['widths'][n]/2)]), (y[i]-pInfo["peak_heights"][n], y[i]-pInfo["peak_heights"][n]))
          plt.annotate(f"X:{xPos:>{size}}\nB:{braggsVal:>{size}}\nA:{area:>{size}}", (x[i], y[i]), xytext=(5,-25), textcoords="offset points", fontfamily="DejaVu Sans Mono", fontsize=8 ,bbox=dict(facecolor='gray', alpha=0.3, edgecolor='none', boxstyle='round'))
  plt.title(f"{year}{number} - {name}\nCrystalline part: {round(crystalRatio, 1)}%", fontweight="bold", fontfamily="arial")
  plt.xlabel("2Θ")
  plt.ylabel("intensity")
  plt.xlim(xMin, xMax)
  plt.ylim(0)
  #plt.legend(loc="upper right")
  plt.gcf().set_size_inches(16, 9)
  
  if testing == True:
    mng = plt.get_current_fig_manager()
    mng.window.state('zoomed')
    plt.show()
  elif testing == False:
    plt.savefig(image, dpi=300)
    plt.close()
  