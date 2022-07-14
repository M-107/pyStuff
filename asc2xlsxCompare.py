import os
import numpy as np
import xlwings as xw
import tkinter as tk
import seaborn as sns
from time import time
from datetime import datetime
from openpyxl import load_workbook
from matplotlib import pyplot as plt 
from shapely.geometry import LineString
from tkinter.filedialog import askopenfilename, askopenfilenames

# Ask for the ASC files and for Excel with formulas
root = tk.Tk()
root.withdraw()
files = askopenfilenames(title="Select the ASC files with RTG data.",
                         filetypes=[("ASC files", ".asc")])
bigExcel = askopenfilename(title="Select the big Excel for calculations.",
                           filetypes=[("Excel files", ".xlsx")])

# Extract the X and Y values from an ASC file
def makeXY(file):
    lines = []
    with open(file, "r") as f:
        for line in f:
            y = line.split()
            lines.append(y)
    lines = lines[:-1]
    x = [e[0] for e in lines]
    y = [f[1] for f in lines]
    return x, y

# Input the X and Y values into the Excel
def excelWrite(excel, x, y):
    wb = load_workbook(excel)
    wsWrite = wb.worksheets[0]
    for i in range(1, 999):
        wsWrite["A" + str(i)].value = None
        wsWrite["B" + str(i)].value = None
    for enum, i in enumerate(x, 1):
        wsWrite["A" + str(enum)].value = i
        wsWrite["B" + str(enum)].value = y[enum-1]
    wb.save(excel)
    wb.close()
    # Open the file in a library that allows recalculating of the formulas inside
    app_excel = xw.App(visible = False)
    wbk = xw.Book(excel)
    wbk.api.RefreshAll()
    wbk.save(excel)
    app_excel.kill()
    del app_excel

# Read the results we're interested in from a recalculated Excel
def excelRead(excel):
    wb = load_workbook(excel, data_only=True)
    wsRead = wb.worksheets[1]
    crystallinePhase = wsRead["I11"].value
    phaseOne = wsRead["I13"].value
    return crystallinePhase, phaseOne

# Group the files into their groups according to the code
# Keeping the naming scheme is important
# CODE NAME melted DD.MM.YYYY HH.MM measured DD.MM.YYYY HH.MM
# The group code is the first part of each file name
def makeGroups(files):
    oneGroup = []
    allGroups = []
    lastCode = ""
    for file in files:
        fileName = os.path.split(file)[1]
        fileSplit = fileName.split(" ")
        code = fileSplit[0]
        nameSplit = fileSplit[1:-6]
        name = ' '.join(str(x) for x in nameSplit)
        meltDate = fileSplit[-5]
        meltTime = fileSplit[-4]
        measureDate = fileSplit[-2]
        measureTime = fileSplit[-1][:-4]
        fileInfo = [file, name, meltDate, meltTime, measureDate, measureTime]
        if code != lastCode:
            allGroups.append(oneGroup)
            oneGroup = [code]
            oneGroup.append(fileInfo)
        else:
            oneGroup.append(fileInfo)
        lastCode = code
    allGroups.append(oneGroup)
    allGroups = allGroups[1:]
    return allGroups

# Get age of each sample during measurement by substracting the time of melting
def getAge(sample):
    meltRaw = " ".join([sample[2], sample[3]])
    measureRaw = " ".join([sample[4], sample[5]])
    melt = datetime.strptime(meltRaw, "%d.%m.%Y %H.%M")
    measure = datetime.strptime(measureRaw, "%d.%m.%Y %H.%M")
    age = measure - melt
    ageSecs = age.total_seconds() 
    ageHours = round(ageSecs/3600, 2)
    return ageHours

# Make a graph showing the relation of phase forms on time
def makeGraph(subList):
    i = subList
    name = i[1][1]
    times = []
    cPhases = []
    phases1 = []
    phases2 = []
    for j in range(1, len(i)):
        timeHrs = i[j][6]
        cPhase = int(round(i[j][7]*100, 2))
        phase1 = int(cPhase * i[j][8])
        phase2 = int(cPhase - phase1)
        times.append(timeHrs)
        cPhases.append(cPhase)
        phases1.append(phase1)
        phases2.append(phase2)
    # Set the visuals
    sns.set_theme()
    sns.set_style("whitegrid")
    plt.plot(times, cPhases, "k", label=f"Crystalline phase")
    plt.plot(times, phases1, "b", label=f"Phase 1")
    plt.plot(times, phases2, "r", label=f"Phase 2")
    plt.xlim(min(times), max(times))
    plt.ylim(0, max(cPhases)+10)
    if name[-1].isnumeric():
        plt.title(f"{name}%")
    else:
        plt.title(name)
    plt.xlabel("Time [h]")
    plt.ylabel("Phase ammount [%]")
    plt.legend(loc="upper left")
    # Find the intersecting point of phase 1 and phase 2 lines if they intersect
    if len(times) > 1 and (phases1[0]-phases2[0])*(phases1[-1]-phases2[-1]) < 0:
        line1 = LineString(np.column_stack((times, phases1)))
        line2 = LineString(np.column_stack((times, phases2)))
        intersection = line1.intersection(line2)
        xInter, yInter = intersection.xy[0][0], intersection.xy[1][0]
        # Plot everything and save it next to the big Excel
        plt.plot(*intersection.xy, 'ko')
        plt.plot([xInter, xInter], [0, yInter], "k--", alpha = 0.4)
        plt.annotate(f"{round(xInter, 2)}", (xInter, yInter), xytext=(-14, 10), 
                     textcoords="offset points",  fontsize=12,
                     bbox=dict(facecolor='white', alpha=0.75, edgecolor='gray', boxstyle='round'))
    plt.gcf().set_size_inches(16, 9)
    plt.savefig(f"{os.path.split(bigExcel)[0]}\{name}.png", dpi=300)
    plt.close()

# Start of the script
print(f"\nWorking on {len(files)} files")
filesGrouped = makeGroups(files)
print(f"Found {len(filesGrouped)} groups:")
for i in filesGrouped:
    for j in i[1:]:
        j.append(getAge(j))

# Print info about the groups
print("-----------------------------------------")
for i in filesGrouped:
    if len(i)-1 != 1:
        word = "files"
    else:
        word = "file"
    print(f"{i[0]:<4}  {i[1][1]:<25} ({len(i)-1} {word})")
print("-----------------------------------------")

# Using the Excel related functions to expand the list of groups
# Then making graphs for each group
tStart = time()
for enum, i in enumerate(filesGrouped):
    t0 = time()
    print(f"\nWorking on group {i[1][1]} ({len(i)-1} files)")
    for j in range(1,len(i)):
        t1 = time()
        x, y = makeXY(i[j][0])
        excelWrite(bigExcel, x, y)
        crystallinePhase, phaseOne = excelRead(bigExcel)
        i[j].append(round(crystallinePhase, 2))
        i[j].append(round(phaseOne,2 ))
        t2 = time()
        print(f"    File {j} took {round(t2-t1, 2):<5} seconds")
    makeGraph(i)
    print(f"The group took {round((t2-t0)/60, 2)} minutes")
tEnd = time()
print(f"\nThe {len(files)} files took {round((tEnd-tStart)/60, 2)} minutes to analyze")
