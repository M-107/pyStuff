import os
import tkinter as tk
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, numbers
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from tkinter.filedialog import askopenfilename, askopenfilenames

# Ask for the working files, prn files are optional
root = tk.Tk()
root.withdraw()
print("Select the cti file(s)")
ctis = askopenfilenames(title="Select the cti file(s)", filetypes=[("cti files", ".cti")])
doPrn = input("Would you like to analyze prn files as well? (Y/N)\n")
if doPrn == "Y" or doPrn == "y":
    print("Select the prn file(s)")
    prns = askopenfilenames(title="Select the prn file(s)", filetypes=[("prn files", ".prn")])

gavePrnErr = 0
for x, ctiFile in enumerate(ctis, 1):
    print(f"\nWorking on cti file {x} out of {len(ctis)}.")
    ctiName = os.path.basename(ctiFile)[:-4]
    # If user wants to analyse prn files as well (doPrn == 1) make sure one with matching name to cti is present
    # Else just act as if doPrn == 0 for this loop (will be managed by foundMatch)
    foundMatch = 0
    if doPrn == "Y" or doPrn == "y":
        for p in prns:
            prnName = os.path.basename(p)[:-8]
            if ctiName == prnName:
                print(f"Found matching prn file")
                prnFile = p
                foundMatch = 1
                break
        if foundMatch == 0:
            print(f"Didn't find matching prn file for {ctiName}. Second excel sheet will not be created.")
            if gavePrnErr == 0:
                print("The names must be EXACTLY the same with prn name having additional '-prn' in the filename.")
                print("For example 'abc123.cti' and 'abc123-prn.prn' would get matched.")
                gavePrnErr = 1

    # Read the cti file as lines
    with open(ctiFile) as f:
        lines = f.read().splitlines() 
    # Find the indices of beggining and ending lines for each data collumn
    freBegin = lines.index("VAR_LIST_BEGIN") + 1
    freEnd = lines.index("VAR_LIST_END")
    s11Begin = lines.index("BEGIN", freEnd) + 1
    s11End = lines.index("END", s11Begin)
    s21Begin = lines.index("BEGIN", s11End) + 1
    s21End = lines.index("END", s21Begin)
    s12Begin = lines.index("BEGIN", s21End) + 1
    s12End = lines.index("END", s12Begin)
    s22Begin = lines.index("BEGIN", s12End) + 1
    s22End = lines.index("END", s22Begin)
    # Put the found data into lists
    # The sXX data have two values on each line, we only care for the first one
    freList = lines[freBegin:freEnd]
    s11List = [i.split(",")[0] for i in lines[s11Begin:s11End]]
    s21List = [i.split(",")[0] for i in lines[s21Begin:s21End]]
    s12List = [i.split(",")[0] for i in lines[s12Begin:s12End]]
    s22List = [i.split(",")[0] for i in lines[s22Begin:s22End]]

    # Read the prn file as lines, if it doesn't exist, ignore this whole step
    if foundMatch == 1:
        with open(prnFile) as f:
            lines = f.read().splitlines() 
        # There might be a different number (n) of collumns with info
        # So make a list of lists (with data for each col) of the length n
        nOfVals = len(lines[3].split())
        prnData = []
        for n in range(0, nOfVals):
            tempList = [i.split()[n] for i in lines[3:]]
            prnData.append(tempList)
        for i in range(0, len(prnData)):
            for j in range(1, len(prnData[0])):
                prnData[i][j] = float(prnData[i][j])
        for i in range(1, len(prnData[0])):
            prnData[0][i] = int(prnData[0][i])

    # Make an excel file and write the static info
    wb = Workbook()
    wsCti = wb.active
    wsCti.title = "cti data"
    wsCti["A1"].value = ctiName
    wsCti["C1"].value = "Odrazí"
    wsCti["D1"].value = "Na druhou"
    wsCti["F1"].value = "Projde"
    wsCti["G1"].value = "Na druhou"
    wsCti["H1"].value = "Absorbuje"
    wsCti["I1"].value = "Na druhou"
    wsCti["J1"].value = "Kontrola"
    wsCti["O1"].value = "SE Total"
    wsCti["P1"].value = "New Power"
    wsCti["R1"].value = "SE Total"
    wsCti["B2"].value = "S11"
    wsCti["C2"].value = "R (%)"
    wsCti["D2"].value = "R2 (%)"
    wsCti["E2"].value = "S21"
    wsCti["F2"].value = "T (%)"
    wsCti["G2"].value = "TR (%)"
    wsCti["H2"].value = "A (%)"
    wsCti["I2"].value = "A2 (%)"
    wsCti["J2"].value = "R2 + T2 + A2"
    wsCti["K2"].value = "S12"
    wsCti["L2"].value = "S22"
    wsCti["M2"].value = "SER"
    wsCti["N2"].value = "SEA"
    wsCti["O2"].value = "SER + SEA"
    wsCti["P2"].value = "SER"
    wsCti["Q2"].value = "SEA"
    wsCti["R2"].value = "SER + SEA"

    # Make the cells wider
    for i in range(1, 19):
        wsCti.column_dimensions[get_column_letter(i)].width = 13
        wsCti[get_column_letter(i) + "2"].border = Border(bottom=Side(border_style=BORDER_THIN, color="00000000"))

    # Colour the cells
    yellow = "FFE699"
    green = "C6E0B4"
    blue = "BDD7EE"
    red = "F8CBAD"
    pink = "FF9999"
    yellowAlert = "FFFF99"
    
    wsCti["A1"].fill = PatternFill(fgColor=yellow, fill_type="solid")
    wsCti["C1"].fill = PatternFill(fgColor=green, fill_type="solid")
    wsCti["D1"].fill = PatternFill(fgColor=green, fill_type="solid")
    wsCti["C2"].fill = PatternFill(fgColor=green, fill_type="solid")
    wsCti["D2"].fill = PatternFill(fgColor=green, fill_type="solid")
    wsCti["F1"].fill = PatternFill(fgColor=blue, fill_type="solid")
    wsCti["G1"].fill = PatternFill(fgColor=blue, fill_type="solid")
    wsCti["F2"].fill = PatternFill(fgColor=blue, fill_type="solid")
    wsCti["G2"].fill = PatternFill(fgColor=blue, fill_type="solid")
    wsCti["H1"].fill = PatternFill(fgColor=red, fill_type="solid")
    wsCti["I1"].fill = PatternFill(fgColor=red, fill_type="solid")
    wsCti["J1"].fill = PatternFill(fgColor=red, fill_type="solid")
    wsCti["H2"].fill = PatternFill(fgColor=red, fill_type="solid")
    wsCti["I2"].fill = PatternFill(fgColor=red, fill_type="solid")
    wsCti["J2"].fill = PatternFill(fgColor=red, fill_type="solid")
    wsCti["P1"].fill = PatternFill(fgColor=pink, fill_type="solid")
    wsCti["Q1"].fill = PatternFill(fgColor=pink, fill_type="solid")
    wsCti["R1"].fill = PatternFill(fgColor=pink, fill_type="solid")
    wsCti["P2"].fill = PatternFill(fgColor=pink, fill_type="solid")
    wsCti["Q2"].fill = PatternFill(fgColor=pink, fill_type="solid")
    wsCti["R2"].fill = PatternFill(fgColor=pink, fill_type="solid")

    # Write cti values
    lastGHz = str(int(freList[0]))[:-9]
    for i in range(0, len(freList)):
        wsCti["A" + str(i + 3)].value = float(freList[i])
        wsCti["B" + str(i + 3)].value = float(s11List[i])
        wsCti["C" + str(i + 3)].value = f"=POWER(10,(B{i + 3}/20))"
        wsCti["D" + str(i + 3)].value = f"=POWER(C{i + 3},2)*(100)"
        wsCti["E" + str(i + 3)].value = float(s21List[i])
        wsCti["F" + str(i + 3)].value = f"=POWER(10,(E{i + 3}/20))"
        wsCti["G" + str(i + 3)].value = f"=POWER(F{i + 3},2)*(100)"
        wsCti["H" + str(i + 3)].value = f"=SQRT(1-((POWER(C{i + 3},2))+(POWER(F{i + 3},2))))"
        wsCti["I" + str(i + 3)].value = f"=POWER(H{i + 3},2)*(100)"
        wsCti["J" + str(i + 3)].value = f"=(D{i + 3}+G{i + 3}+I{i + 3})"
        wsCti["K" + str(i + 3)].value = float(s12List[i])
        wsCti["L" + str(i + 3)].value = float(s22List[i])
        wsCti["M" + str(i + 3)].value = f"=ABS(10*LOG(1/ABS(1-B{i + 3}^2)))"
        wsCti["N" + str(i + 3)].value = f"=ABS(10*LOG(ABS((1-B{i + 3}^2)/K{i + 3}^2)))"
        wsCti["O" + str(i + 3)].value = f"=ABS(M{i + 3}+N{i + 3})"
        wsCti["P" + str(i + 3)].value = f"=10*LOG(1/(1-10^(B{i + 3}/10)))"
        wsCti["Q" + str(i + 3)].value = f"=10*LOG((1-10^(B{i + 3}/10))/10^(E{i + 3}/10))"
        wsCti["R" + str(i + 3)].value = f"=ABS(P{i + 3}+Q{i + 3})"
        # Colour the cells for each collumn that needs it
        wsCti["C" + str(i + 3)].fill = PatternFill(fgColor=green, fill_type="solid")
        wsCti["D" + str(i + 3)].fill = PatternFill(fgColor=green, fill_type="solid")
        wsCti["F" + str(i + 3)].fill = PatternFill(fgColor=blue, fill_type="solid")
        wsCti["G" + str(i + 3)].fill = PatternFill(fgColor=blue, fill_type="solid")
        wsCti["H" + str(i + 3)].fill = PatternFill(fgColor=red, fill_type="solid")
        wsCti["I" + str(i + 3)].fill = PatternFill(fgColor=red, fill_type="solid")
        wsCti["J" + str(i + 3)].fill = PatternFill(fgColor=red, fill_type="solid")
        wsCti["P" + str(i + 3)].fill = PatternFill(fgColor=pink, fill_type="solid")
        wsCti["q" + str(i + 3)].fill = PatternFill(fgColor=pink, fill_type="solid")
        wsCti["r" + str(i + 3)].fill = PatternFill(fgColor=pink, fill_type="solid")
        # Yellow the line when frequency GHz changes
        GHz = str(int(freList[i]))[:-9]
        if GHz != lastGHz:
            for j in range(1,19):
                wsCti[get_column_letter(j) + str(i + 3)].fill = PatternFill(fgColor=yellowAlert, fill_type="solid")
        lastGHz = GHz
    
    # Open the prn workseet if prn file was foudn and write the prn values
    # Also do all the cell eiditng like in cti worksheet
    if foundMatch == 1:
        wsPrn = wb.create_sheet("prn data")
        wsPrn["A1"].value = ctiName
        wsPrn["A1"].fill = PatternFill(fgColor=yellow, fill_type="solid")
        for i in range(1, 7):
            wsPrn.column_dimensions[get_column_letter(i)].width = 13
            wsPrn[get_column_letter(i) + "2"].border = Border(bottom=Side(border_style=BORDER_THIN, color="00000000"))
        for i in range(0, len(prnData)):
            for j in range(0, len(prnData[0])):
                wsPrn[get_column_letter(i + 2) + str(j + 2)].value = prnData[i][j]
    
    # Save and close the excel file
    wb.save(ctiFile[:-4] + ".xlsx")
    wb.close()

print("\nAll done\nPress any key to exit")
input()