import os
import re
import tkinter as tk
from time import sleep, time
from pyautogui import hotkey
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from tkinter.filedialog import askdirectory, askopenfilename

# Hide the tk window
root = tk.Tk()
root.withdraw()

# Get Excel and ASC files
excelFirst = askopenfilename(title = "Select Excel file for calculations.") # This is the bigbrain one
excelSecond = askopenfilename(title = "Select Excel file for results.") 
# Second Excel can be empty, first col is left blank for value descriptions
# TODO Ged rid of the empy sheet, replace with autoamtically making one with correct col names
path = askdirectory(title = "Select folder with ASC files.") 

# Make and prepare the results excel file


# Iterate through all ASC files in path and subdirs
print(f"Found {len(os.listdir(path))} files.")
for x, dataFile in enumerate(os.listdir(path), 1):
    t1 = time()
    print(f"Working on file {x} out of {len(os.listdir(path))}.")
    
    # Group 1 will be name, 2 melt time, 3 measure time
    # This should find melt and measurement info even in badly named ASC files
    findImportant = re.search("(.*), tav.\s?(\d*\.\d*\s\d*\.\d*\.\d*).*mer.\s?(\d*\.\d*\.\d*\s\d*\.\d*)", dataFile)
    name = findImportant.group(1)
    # Get first datetime group, split it by spaces, first split is date, split that by dots to get d, m, y
    tDateRaw = findImportant.group(2).split()[1].split(".") 
    tTimeRaw = findImportant.group(2).split()[0].split(".") # Same, but with time
    # Minutes of melt and measurement sometimes aren't written in the name, they aren't THAT important, so just replace them with "00"
    if not tTimeRaw[1]:
        tTimeRaw[1] = "00"
        print("Melt time didn't contain minutes, inputing '00'.")
    tDate = f"{int(tDateRaw[0]):02}.{int(tDateRaw[1]):02}.{int(tDateRaw[2]):04}"
    tTime = f"{int(tTimeRaw[0]):02}:{int(tTimeRaw[1]):02}"
    # Melt and measurement date and time are written in different order, that's why the splits are switched
    mDateRaw = findImportant.group(3).split()[0].split(".") 
    mTimeRaw = findImportant.group(3).split()[1].split(".")
    if not mTimeRaw[1]:
        mTimeRaw[1] = "00"
        print("Measurement time didn't contain minutes, inputing '00'.")
    mDate = f"{int(mDateRaw[0]):02}.{int(mDateRaw[1]):02}.{int(mDateRaw[2]):04}"
    mTime = f"{int(mTimeRaw[0]):02}:{int(mTimeRaw[1]):02}"
    
    # Load calculating Excel and input the name and datetime info
    wb = load_workbook(excelFirst)
    wsWrite = wb.worksheets[0]
    wsWrite["D14"].value = name
    wsWrite["D15"].value = tDate
    wsWrite["D16"].value = tTime
    wsWrite["D17"].value = mDate
    wsWrite["D18"].value = mTime
    print(f"Name:{name}   Melting:{tDate} {tTime}  Measuring:{mDate} {mTime}")
    
    # Clear previous data from working cells
    for i in range(1, 999):
        wsWrite["A" + str(i)].value = None
        wsWrite["B" + str(i)].value = None
    
    # Read the ASC file, convert numbers to Excel-friendly format, input them in the working cells
    text = open(os.path.join(path, dataFile), mode="r", encoding="cp1250")
    lines = text.readlines()
    for y, j in enumerate(lines, 1):
        left = j[1:17]
        if left == "":
            break
        exponent = int(j[22])
        right = j[25:len(j)+1]
        wsWrite["A" + str(y)].value = float(left) * 10 ** exponent
        wsWrite["B" + str(y)].value = int(right)
    wb.save(excelFirst)
    wb.close()
    
    # Actually open the Excel file, let it recalculate and save it. DO NOT TOUCH THE PC during this step!
    # The file must be loaded, no other way aroudn this apart from doing all the calulations in Python
    # Which would make the Excel file obsolete and me depressed
    print("File is updated and saved, opening Excel to recalculate.")
    print("   Don't press anything during this step.")
    os.startfile(excelFirst, "open")
    sleep(30)
    hotkey("ctrl", "s")
    sleep(20)
    hotkey("alt", "f4")
    sleep(10)
    
    # Load the recalculated Excel file
    print("Data recalculated, opening file again to read new data.")
    wbRead = load_workbook(excelFirst, data_only=True)
    wsRead = wbRead.worksheets[1]
    valuesNew = []
    for i in range(2, 61):
        value = wsRead["I" + str(i)].value
        if value != None:
            valuesNew.append(value)
    wbRead.close()
    
    # Write the results to a second Excel file, new collumn every loop
    wbWrite = load_workbook(excelSecond, read_only=False)
    wsWrite = wbWrite.worksheets[0]
    for i in range(0, len(valuesNew)):
        wsWrite[get_column_letter(x + 1) + str(i + 1)].value = valuesNew[i]
    wbWrite.save(excelSecond)
    wbWrite.close()
    print("New data saved in second Excel file.")

    # Write values from each loop into the second excel file
    
    
    # Check the loop duration, might need increasing of sleeps on slow PCs
    t2 = time()
    print(f"This loop took {round(t2 - t1, 2)} seconds.\n")